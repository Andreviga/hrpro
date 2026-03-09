import json
import openpyxl


def to_serializable(value):
    if value is None:
        return None
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)

path = r"d:/HRPro Folha de Pagamento/Folha de pagamento de fevereiro 2026.xlsm"
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)

focus_sheets = [
    "Cadastro Funcionários",
    "Folha de pagto janeiro2026",
    "Folha de pagto 13 2025",
    "Folha de pagto Férias ",
    "Holerite",
    "TRCT",
    "VT012026",
    "Recibo de Férias",
    "Quantidade de aula 0125",
    "Tab auxílio 0125",
    "Folha de pagto extra 0125",
]

result = {}

for sheet_name in focus_sheets:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    header_preview = []
    for r in range(1, 6):
        row_vals = []
        for c in range(1, min(21, ws.max_column + 1)):
            val = ws.cell(row=r, column=c).value
            row_vals.append(to_serializable(val))
        header_preview.append(row_vals)

    ref_errors = []
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.startswith("=") and "#REF!" in val:
                ref_errors.append({"cell": cell.coordinate, "formula": val})
            elif cell.data_type == "f" and val and "#REF!" in str(val):
                ref_errors.append({"cell": cell.coordinate, "formula": str(val)})

    result[sheet_name] = {
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "header_preview": header_preview,
        "ref_errors": ref_errors,
    }

print(json.dumps(result, ensure_ascii=True, indent=2))
