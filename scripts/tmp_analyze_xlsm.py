import json
import openpyxl

path = r"d:/HRPro Folha de Pagamento/Folha de pagamento de fevereiro 2026.xlsm"
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)

summary = []
total_ref = 0

for ws in wb.worksheets:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    formula_cells = []
    ref_cells = []

    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formula_cells.append(cell.coordinate)
                if "#REF!" in value:
                    ref_cells.append(cell.coordinate)
            elif cell.data_type == "f" and value:
                formula_cells.append(cell.coordinate)
                if "#REF!" in str(value):
                    ref_cells.append(cell.coordinate)

    total_ref += len(ref_cells)
    summary.append(
        {
            "sheet": ws.title,
            "max_row": max_row,
            "max_col": max_col,
            "formula_count": len(formula_cells),
            "ref_error_count": len(ref_cells),
            "ref_error_cells": ref_cells[:20],
        }
    )

print(
    json.dumps(
        {"sheets": summary, "total_ref_errors": total_ref},
        ensure_ascii=True,
        indent=2,
    )
)
